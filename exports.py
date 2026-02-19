import io
from datetime import date

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from flask import Blueprint, request, send_file
from flask_login import login_required
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT

from models import Aviso, ESTADOS
from extensions import db

exports_bp = Blueprint('exports', __name__, url_prefix='/export')

ESTADO_LABELS = {key: label for key, label in ESTADOS}


@exports_bp.route('/excel')
@login_required
def export_excel():
    estado_filter = request.args.get('estado', '')
    q = request.args.get('q', '').strip()

    query = Aviso.query

    if estado_filter:
        query = query.filter_by(estado=estado_filter)

    if q:
        like = f'%{q}%'
        query = query.filter(
            db.or_(
                Aviso.nombre_cliente.ilike(like),
                Aviso.telefono.ilike(like),
                Aviso.calle.ilike(like),
            )
        )

    avisos = query.order_by(Aviso.fecha_aviso.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Avisos'

    headers = ['ID', 'Fecha Aviso', 'Fecha Cita', 'Cliente', 'Teléfono',
               'Calle', 'Localidad', 'Electrodoméstico', 'Marca',
               'Descripción', 'Estado', 'Notas']

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(fill_type='solid', fgColor='1a6496')

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for row, aviso in enumerate(avisos, 2):
        ws.cell(row=row, column=1, value=aviso.id)
        ws.cell(row=row, column=2,
                value=aviso.fecha_aviso.strftime('%d/%m/%Y') if aviso.fecha_aviso else '')
        ws.cell(row=row, column=3,
                value=aviso.fecha_cita.strftime('%d/%m/%Y') if aviso.fecha_cita else '')
        ws.cell(row=row, column=4, value=aviso.nombre_cliente)
        ws.cell(row=row, column=5, value=aviso.telefono)
        ws.cell(row=row, column=6, value=aviso.calle or '')
        ws.cell(row=row, column=7, value=aviso.localidad or '')
        ws.cell(row=row, column=8, value=aviso.electrodomestico or '')
        ws.cell(row=row, column=9, value=aviso.marca or '')
        ws.cell(row=row, column=10, value=aviso.descripcion or '')
        ws.cell(row=row, column=11, value=ESTADO_LABELS.get(aviso.estado, aviso.estado))
        ws.cell(row=row, column=12, value=aviso.notas or '')

    # Auto-ajustar ancho de columnas
    for col in ws.columns:
        max_len = max((len(str(cell.value or '')) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    suffix = f'_{estado_filter}' if estado_filter else ''
    filename = f'avisos{suffix}_{date.today().strftime("%Y%m%d")}.xlsx'

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename,
    )


# ── Albarán PDF ────────────────────────────────────────────────────────────

@exports_bp.route('/albaran/<int:id>')
@login_required
def albaran_pdf(id):
    aviso = Aviso.query.get_or_404(id)

    output = io.BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=A4,
        leftMargin=2*cm, rightMargin=2*cm,
        topMargin=2*cm, bottomMargin=2*cm,
    )

    styles = getSampleStyleSheet()
    azul   = colors.HexColor('#1a6496')
    gris   = colors.HexColor('#666666')
    verde  = colors.HexColor('#28a745')
    rojo   = colors.HexColor('#dc3545')

    estilo_titulo   = ParagraphStyle('titulo',   fontSize=22, textColor=azul,  spaceAfter=2,  fontName='Helvetica-Bold')
    estilo_empresa  = ParagraphStyle('empresa',  fontSize=10, textColor=gris,  spaceAfter=1,  fontName='Helvetica')
    estilo_h2       = ParagraphStyle('h2',       fontSize=12, textColor=azul,  spaceBefore=12, spaceAfter=4, fontName='Helvetica-Bold')
    estilo_normal   = ParagraphStyle('normal',   fontSize=10, spaceAfter=2,    fontName='Helvetica')
    estilo_negrita  = ParagraphStyle('negrita',  fontSize=10, fontName='Helvetica-Bold')
    estilo_muted    = ParagraphStyle('muted',    fontSize=9,  textColor=gris,  fontName='Helvetica-Oblique')
    estilo_total    = ParagraphStyle('total',    fontSize=13, fontName='Helvetica-Bold', textColor=azul, alignment=TA_RIGHT)
    estilo_beneficio = ParagraphStyle('beneficio', fontSize=11, fontName='Helvetica-Bold', alignment=TA_RIGHT)

    elementos = []

    # ── Cabecera ──
    elementos.append(Paragraph('🔧 CadizTécnico', estilo_titulo))
    elementos.append(Paragraph('Servicio técnico de electrodomésticos', estilo_empresa))
    elementos.append(HRFlowable(width='100%', thickness=2, color=azul, spaceAfter=10))

    # ── Número y fecha ──
    fecha_str = date.today().strftime('%d/%m/%Y')
    tabla_ref = Table(
        [[Paragraph(f'<b>ALBARÁN Nº {aviso.id:04d}</b>', ParagraphStyle('ref', fontSize=14, fontName='Helvetica-Bold')),
          Paragraph(f'Fecha: {fecha_str}', ParagraphStyle('fecha', fontSize=10, alignment=TA_RIGHT, fontName='Helvetica'))]],
        colWidths=[10*cm, 7*cm]
    )
    tabla_ref.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 8),
    ]))
    elementos.append(tabla_ref)
    elementos.append(Spacer(1, 0.3*cm))

    # ── Datos del cliente ──
    elementos.append(Paragraph('DATOS DEL CLIENTE', estilo_h2))
    datos_cliente = [
        ['Nombre:', aviso.nombre_cliente],
        ['Teléfono:', aviso.telefono],
    ]
    if aviso.calle:
        dir_ = aviso.calle + (f', {aviso.localidad}' if aviso.localidad else '')
        datos_cliente.append(['Dirección:', dir_])
    if aviso.fecha_cita:
        datos_cliente.append(['Fecha visita:', aviso.fecha_cita.strftime('%d/%m/%Y')])

    tabla_cliente = Table(datos_cliente, colWidths=[4*cm, 13*cm])
    tabla_cliente.setStyle(TableStyle([
        ('FONTNAME',  (0,0), (0,-1), 'Helvetica-Bold'),
        ('FONTNAME',  (1,0), (1,-1), 'Helvetica'),
        ('FONTSIZE',  (0,0), (-1,-1), 10),
        ('TEXTCOLOR', (0,0), (0,-1), gris),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
    ]))
    elementos.append(tabla_cliente)

    # ── Trabajo realizado ──
    elementos.append(Paragraph('TRABAJO REALIZADO', estilo_h2))
    datos_trabajo = [
        ['Aparato:', f'{aviso.electrodomestico or "—"}  {("· " + aviso.marca) if aviso.marca else ""}'],
    ]
    if aviso.descripcion:
        datos_trabajo.append(['Avería:', aviso.descripcion])
    if aviso.notas:
        datos_trabajo.append(['Notas:', aviso.notas])

    tabla_trabajo = Table(datos_trabajo, colWidths=[4*cm, 13*cm])
    tabla_trabajo.setStyle(TableStyle([
        ('FONTNAME',  (0,0), (0,-1), 'Helvetica-Bold'),
        ('FONTNAME',  (1,0), (1,-1), 'Helvetica'),
        ('FONTSIZE',  (0,0), (-1,-1), 10),
        ('TEXTCOLOR', (0,0), (0,-1), gris),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
    ]))
    elementos.append(tabla_trabajo)

    # ── Materiales empleados ──
    if aviso.materiales_desc:
        elementos.append(Paragraph('MATERIALES / PIEZAS', estilo_h2))
        elementos.append(Paragraph(aviso.materiales_desc, estilo_normal))

    # ── Resumen económico ──
    tiene_precio = aviso.precio_mano_obra is not None
    tiene_coste  = aviso.coste_materiales is not None

    if tiene_precio or tiene_coste:
        elementos.append(Spacer(1, 0.5*cm))
        elementos.append(HRFlowable(width='100%', thickness=1, color=colors.lightgrey))
        elementos.append(Spacer(1, 0.3*cm))

        filas_eco = []
        if tiene_precio:
            filas_eco.append([
                Paragraph('Mano de obra:', ParagraphStyle('eco_lab', fontSize=10, fontName='Helvetica', textColor=gris)),
                Paragraph(f'{aviso.precio_mano_obra:.2f} €', ParagraphStyle('eco_val', fontSize=10, fontName='Helvetica', alignment=TA_RIGHT)),
            ])
        if tiene_coste:
            filas_eco.append([
                Paragraph('Materiales:', ParagraphStyle('eco_lab', fontSize=10, fontName='Helvetica', textColor=gris)),
                Paragraph(f'{aviso.coste_materiales:.2f} €', ParagraphStyle('eco_val', fontSize=10, fontName='Helvetica', alignment=TA_RIGHT)),
            ])

        total = (aviso.precio_mano_obra or 0) + (aviso.coste_materiales or 0)
        filas_eco.append([
            Paragraph('<b>TOTAL:</b>', ParagraphStyle('total_lab', fontSize=13, fontName='Helvetica-Bold')),
            Paragraph(f'<b>{total:.2f} €</b>', ParagraphStyle('total_val', fontSize=13, fontName='Helvetica-Bold', alignment=TA_RIGHT, textColor=azul)),
        ])

        tabla_eco = Table(filas_eco, colWidths=[13*cm, 4*cm])
        tabla_eco.setStyle(TableStyle([
            ('LINEABOVE',  (0, -1), (-1, -1), 1.5, azul),
            ('TOPPADDING', (0, -1), (-1, -1), 6),
            ('BOTTOMPADDING', (0,0), (-1,-1), 4),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ]))
        elementos.append(tabla_eco)

    # ── Pie de página ──
    elementos.append(Spacer(1, 1*cm))
    elementos.append(HRFlowable(width='100%', thickness=1, color=colors.lightgrey))
    elementos.append(Spacer(1, 0.3*cm))
    elementos.append(Paragraph(
        'Gracias por confiar en CadizTécnico · Garantía de 3 meses en reparaciones',
        ParagraphStyle('pie', fontSize=8, textColor=gris, alignment=TA_CENTER, fontName='Helvetica-Oblique')
    ))

    doc.build(elementos)
    output.seek(0)

    nombre_pdf = f'albaran_{aviso.id:04d}_{aviso.nombre_cliente.replace(" ","_")}.pdf'
    return send_file(output, mimetype='application/pdf',
                     as_attachment=False, download_name=nombre_pdf)
